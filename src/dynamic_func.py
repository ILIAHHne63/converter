import openpyxl
import requests
from bs4 import BeautifulSoup
import src.Globals as G
import datetime

def get_exchange_list(currency_from_to, date):
    """Returns res_list and date_list to create a plot"""
    wb = openpyxl.reader.excel.load_workbook(filename="src/dynamic.xlsx")
    wb.active = 0
    sheet = wb.active
    xl_currency_from_to = [-1, -1]
    exchange_from_to_list = [[], []]
    for i in range(len(currency_from_to)):
        xl_currency_from_to[i] = is_in_xl(currency_from_to[i], date)
        if xl_currency_from_to[i] == 0:
            exchange_from_to_list[i] = parce_exchange(currency_from_to[i], date)
        else:
            exchange_from_to_list[i] = get_xl(xl_currency_from_to[i])
        if exchange_from_to_list[i] == -1:
            return -1, -1
    print(exchange_from_to_list[0], len(exchange_from_to_list[0]))
    print(exchange_from_to_list[1], len(exchange_from_to_list[1]))

    data_list = get_date_list()
    print(data_list, len(data_list))
    res_list = [float(exchange_from_to_list[0][i].replace(",", ".")) / float(exchange_from_to_list[1][i].replace(",", ".")) for i in range(G.DATE_COUNTER-1)]
    return data_list, res_list


def is_in_xl(currency, data):
    """If currency exist in excel returns index else return zero"""
    wb = openpyxl.reader.excel.load_workbook(filename="src/dynamic.xlsx")
    wb.active = 0
    sheet = wb.active
    if sheet["A1"].value is None:
        sheet["A1"].value = 0
        wb.save("src/dynamic.xlsx")
        wb.close()
        return 0
    for i in range(2, sheet["A1"].value+2):
        if sheet.cell(row=1, column=i).value == currency:
            k = 2
            while not(sheet.cell(row=k, column=i).value is None):
                k += 1
            G.DATE_COUNTER = k - 1
            return i
    return 0


def parce_exchange(currency, date=1):
    """A part of the parcing; Returns Currency list in RUBles"""
    wb = openpyxl.reader.excel.load_workbook(filename="src/dynamic.xlsx")
    wb.active = 0
    sheet = wb.active
    if sheet["A1"].value is None:
        sheet["A1"].value = 0
    value = check_is_rub(currency)
    rus_flag = 0
    if value == -1:
        if G.DATE_COUNTER != 0:
            return ["1.0" for i in range(G.DATE_COUNTER)]
        rus_flag = 1
        value = G.code_Belarus_rubl #code of BYN currency
        currency = "BYN"

    soup = init_parc(date, value)
    trs = parce_soup(soup)
    if trs == -1:
        return -1
    date_list, rub_list = parce_trs(trs, sheet, currency)
    sheet["A1"].value += 1
    wb.save("src/dynamic.xlsx")
    wb.close()
    if rus_flag:
        return ["1.0" for i in range(G.DATE_COUNTER)]
    return rub_list


def get_xl(index):
    """Returns exchange rate using index"""
    wb = openpyxl.reader.excel.load_workbook(filename="src/dynamic.xlsx")
    wb.active = 0
    sheet = wb.active
    currency_list = []
    for j in range(2, G.DATE_COUNTER + 1):
        currency_list.append(sheet.cell(row=j, column=index).value)
    return currency_list


def get_date_list():
    """Return Date_list"""
    wb = openpyxl.reader.excel.load_workbook(filename="src/dynamic.xlsx")
    wb.active = 0
    sheet = wb.active
    data_list = []
    for i in range(2, G.DATE_COUNTER + 1):
            data_list.append(sheet.cell(row=i, column=1).value[:5])
    return data_list


def parce_trs(trs, sheet, currency):
    """A part of the parcing; Return date_list and currencu list in RUBles"""
    date_list = []
    rub_list = []
    counter = 2
    G.DATE_COUNTER = 1
    for tr in trs:
        tds = tr.find_all("td")
        if tds is None:
            return -1
        if len(tds) == 3:
            date_list.append(tds[0].text)
            rub_list.append(tds[2].text)
            G.DATE_COUNTER += 1
            sheet["A" + str(counter)].value = tds[0].text
            sheet.cell(row=1, column=sheet["A1"].value+2).value = currency
            sheet.cell(row=counter, column=sheet["A1"].value+2).value = tds[2].text
            counter += 1
    print(rub_list,"rub_list")
    return date_list, rub_list


def parce_soup(soup):
    """A part of the parcing; returns trs ;or -1 if anything is bad"""
    if soup is None:
        return -1
    l = soup.find("table", class_="data")
    if l is None:
        return -1
    trs = l.find_all("tr")
    if trs is None or trs == []:
        return -1
    return trs


def check_is_rub(currency):
    """Checks if currency is RUBles"""
    value = -1
    for list in G.cur:
        if list[1] == currency:
            value = list[2]
    return value



def init_parc(date1, val):
    """Initializing of parcing; Returns soup object"""
    days_to_subtract = 30
    d1 = datetime.datetime.today()
    d2 = d1 - datetime.timedelta(days=days_to_subtract)
    d1 = d1.strftime("%d.%m.%Y")
    d2 = d2.strftime("%d.%m.%Y")
    print(d1, d2)
    url = f"https://www.cbr.ru/currency_base/dynamics/?UniDbQuery.Posted=True&UniDbQuery.so=0&UniDbQuery.mode=1&UniDbQuery.date_req1=&UniDbQuery.date_req2=&UniDbQuery.VAL_NM_RQ=R0{val}&UniDbQuery.From="+str(d2)+"&UniDbQuery.To="+str(d1)
    headers = {
        "Accept": "*/*",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.54 Safari/537.36 Edg/101.0.1210.39"
    }
    print(url)
    req = requests.get(url, headers=headers)
    src = req.text
    soup = BeautifulSoup(src, "lxml")
    return soup
